import datetime
import requests
import json
import openpyxl


Get_China=r"https://view.inews.qq.com/g2/getOnsInfo?name=disease_h5"

class item:

    def __init__(self):
        self.country=list()#国家
        self.province = list()#省份
        self.area=list()#地区
        self.confirm=list()#确诊
        self.suspect=list()#疑似
        self.heal=list()#治愈
        self.dead=list()#死亡
Data_Box=item()#数据盒子

def GetHtmlText(url):
    try:
        res = requests.get(url,timeout = 30)
        res.raise_for_status()
        res.encoding = res.apparent_encoding
        return res.text
    except:

        return "Error"
#获取Json
China = GetHtmlText(Get_China)
City_Count_json = json.loads(China)
City_Count_json = City_Count_json["data"]                   #将json数据中的data字段的数据提取处理
City_Count_json = json.loads(City_Count_json)               #将提取出的字符串转换为json数据

#获取每日总信息
lastUpdateTime = City_Count_json["lastUpdateTime"]
chinaTotal_json = City_Count_json["chinaTotal"]            #提取处其chinaTotal字段中的数据
confirmCount = str(chinaTotal_json["confirm"])
suspectCount = str(chinaTotal_json["suspect"])
deadCount = str(chinaTotal_json["dead"])                   #GetTextCenter(China,r"\"deadCount\": ",r",\n")      #疑似人数
cure = str(chinaTotal_json["heal"])                             #GetTextCenter(China,r"\"cure\": ",r"\n")       #治愈人数
print("更新时间：" + lastUpdateTime + "\n" + "确诊人数为：" + confirmCount + "人\n" + "死亡人数为：" +
      deadCount + "人\n" + "疑似人数为：" + suspectCount + "人\n" + "治愈人数为：" + cure +
      "人\n" )

'''
用于循环中备注信息，防止混淆变量名而出错，然而还是耗费了相当长的时间理清这些变量
areaTree_json[i]["children"]省份
areaTree_json[i]["children"][j]["name"]省份名
areaTree_json[i]["children"][j]["children"][n]省份中的地区 list
areaTree_json[i]["children"][j]["children"][n]省份中的地区 json
areaTree_json[i]["children"][j]["children"][n]["name"]省份中的地区名
areaTree_json[i]["children"][j]["children"][n]["total"]省份中的地区数据json {'confirm': 134, 'suspect': 0, 'dead': 0, 'heal': 4}
'''

areaTree_json=City_Count_json["areaTree"]#包含国家、省份、地区的所有信息，且国家为首索引

def Get_Data_China():
    country_len = len(areaTree_json)
    for i in range(0,country_len):
        if(areaTree_json[i]["name"]=="中国"):            #如果为中国则说明具有省份信息
            province_len = len(areaTree_json[i]["children"])  #获取省份长度
            for j in range(0,province_len):
                area_len=len(areaTree_json[i]["children"][j]["children"])#获取地区长度
                for n in range(0,area_len):
                    total=areaTree_json[i]["children"][j]["children"][n]["total"]                  #获取地区的总体疫情情况+
                    Data_Box.country.append("中国")
                    Data_Box.province.append(areaTree_json[i]["children"][j]["name"])
                    Data_Box.area.append(areaTree_json[i]["children"][j]["children"][n]["name"])
                    Data_Box.confirm.append(total["confirm"])
                    Data_Box.dead.append(total["dead"])
                    Data_Box.heal.append(total["heal"])  #中国区域获取完毕
        else:#外国区域
            name=areaTree_json[i]["name"]
            total=areaTree_json[i]["total"]
            Data_Box.country.append(name)
            Data_Box.province.append(name)
            Data_Box.area.append(name)
            Data_Box.confirm.append(total["confirm"])
            Data_Box.suspect.append(total["suspect"])
            Data_Box.dead.append(total["dead"])
            Data_Box.heal.append(total["heal"])  #外国区域获取完毕
    return len(Data_Box.area)
length=Get_Data_China()#获取信息并获取长度
print("国家  省份  地区  确诊人数  治愈人数  死亡人数  ")

for n in range(0,length):
    print(Data_Box.country[n]+"  "+Data_Box.province[n]+"  "+Data_Box.area[n]+"  "
          +str(Data_Box.confirm[n])+"  "+str(Data_Box.heal[n])+"  "+ str(Data_Box.dead[n]))

def write(length):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, value="国家")
    ws.cell(1, 2, value="省份")
    ws.cell(1, 3, value="地区")
    ws.cell(1, 4, value="确诊人数")
    ws.cell(1, 5, value="治愈人数")
    ws.cell(1, 6, value="死亡人数")
    for n in range(0,length):
        ws.cell(n + 2, 1, Data_Box.country[n])
        ws.cell(n + 2, 2, Data_Box.province[n])
        ws.cell(n + 2, 3, Data_Box.area[n])
        ws.cell(n + 2, 4, Data_Box.confirm[n])
        ws.cell(n + 2, 5, Data_Box.heal[n])
        ws.cell(n + 2, 6, Data_Box.dead[n])
    now = datetime.datetime.now()
    month, day = now.month,now.day
    wb.save(str(month)+str('.')+str(day)+" "+"data_new.xlsx")
    return


write(length)
